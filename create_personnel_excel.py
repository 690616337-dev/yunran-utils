#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "人员信息统计"

# 设置表头
headers = ["序号", "姓名", "身份证号码", "备注"]
ws.append(headers)

# 设置表头样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
header_alignment = Alignment(horizontal="center", vertical="center")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# 添加示例数据（空行供填写）
for i in range(2, 22):  # 20行空白数据行
    ws.append([i-1, "", "", ""])

# 设置列宽
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 20

# 设置边框
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=1, max_row=21, min_col=1, max_col=4):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

# 保存文件
output_path = "/root/.openclaw/workspace/人员信息统计表.xlsx"
wb.save(output_path)
print(f"文件已创建: {output_path}")
