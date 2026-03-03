#!/usr/bin/env python3
"""
生成2026年体育舞蹈赛事Excel文件
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel():
    # 读取JSON数据
    with open('/root/.openclaw/workspace/2026年体育舞蹈赛事数据库.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    competitions = data['competitions']
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "2026年赛事列表"
    
    # 设置标题
    ws['A1'] = "2026年体育舞蹈赛事数据库"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 30
    
    # 设置表头
    headers = ["日期", "地点", "比赛名称", "数据渠道"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 设置边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 填充数据
    for row, comp in enumerate(competitions, 3):
        ws.cell(row=row, column=1, value=comp['date'])
        ws.cell(row=row, column=2, value=comp['city'])
        ws.cell(row=row, column=3, value=comp['name'])
        ws.cell(row=row, column=4, value=comp['source'])
        
        # 设置边框和对齐
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # 四川地区高亮
            if '四川' in comp['city'] or '成都' in comp['city']:
                cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 25
    
    # 冻结首行
    ws.freeze_panes = 'A3'
    
    # 保存文件
    excel_path = '/root/.openclaw/workspace/2026年体育舞蹈赛事数据库.xlsx'
    wb.save(excel_path)
    
    print(f"Excel文件已生成: {excel_path}")
    print(f"共 {len(competitions)} 场赛事")
    
    # 统计信息
    print("\n数据来源统计:")
    source_count = {}
    for comp in competitions:
        source = comp['source']
        source_count[source] = source_count.get(source, 0) + 1
    for source, count in sorted(source_count.items(), key=lambda x: -x[1]):
        print(f"  {source}: {count}场")
    
    return excel_path

if __name__ == '__main__':
    generate_excel()
